

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import calendar
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.filters import OrderingFilter, SearchFilter
from django.core.files.storage import default_storage
from django_filters.rest_framework import DjangoFilterBackend
import math

from .models import Employee, Department, AttendanceRecord, EmployeeDepartment, Salary, Holiday, Configuration
from .serializers import EmployeeSerializer, AttendanceRecordSerializer, SalarySerializer, DepartmentSerializer, \
    EmployeeDepartmentSerializer, HolidaySerializer, ConfigurationSerializer


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = ['name', 'employee_code', 'departments__name']
    ordering_fields = ['name', 'employee_code', 'date_of_birth']
    search_fields = ['name', 'employee_code', 'phone_number', 'date_of_birth']
    # search_fields = ['employee_code']
    ordering = ['name']  # Sắp xếp mặc định


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer


class EmployeeDepartmentViewSet(viewsets.ModelViewSet):
    queryset = EmployeeDepartment.objects.all()
    serializer_class = EmployeeDepartmentSerializer


class AttendanceRecordViewSet(viewsets.ModelViewSet):
    queryset = AttendanceRecord.objects.all()
    serializer_class = AttendanceRecordSerializer


# class SalaryViewSet(viewsets.ModelViewSet):
#     queryset = Salary.objects.all()
#     serializer_class = SalarySerializer
#     filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
#     filterset_fields = ['id', 'employee', 'month', 'year', 'total_hours', 'salary_amount']
#     ordering_fields = ['employee', 'month', 'year', 'total_hours', 'salary_amount']
#     search_fields = ['employee__name', 'employee__employee_code', 'month', 'year']
#     ordering = ['employee']

from django_filters import rest_framework as filters

class SalaryFilter(filters.FilterSet):
    department = filters.NumberFilter(field_name='employee__departments__id')

    class Meta:
        model = Salary
        fields = ['id', 'department', 'employee', 'month', 'year']


class SalaryViewSet(viewsets.ModelViewSet):
    queryset = Salary.objects.all()
    serializer_class = SalarySerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_class = SalaryFilter
    ordering_fields = ['employee', 'month', 'year', 'total_hours', 'salary_amount']
    search_fields = ['employee__name', 'employee__employee_code', 'month', 'year']
    ordering = ['employee']



class HolidayViewSet(viewsets.ModelViewSet):
    queryset = Holiday.objects.all()
    serializer_class = HolidaySerializer


class ConfigurationViewSet(viewsets.ModelViewSet):
    queryset = Configuration.objects.all()
    serializer_class = ConfigurationSerializer



class EmployeeMonthlyAttendanceView(APIView):
    def get(self, request, *args, **kwargs):
        employee_code = request.query_params.get('employee_code')
        month = request.query_params.get('month')
        year = request.query_params.get('year')

        if not employee_code or not month or not year:
            return Response({'error': 'Employee code, month, and year are required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(employee_code=employee_code)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        attendance_records = AttendanceRecord.objects.filter(
            employee=employee,
            date__year=year,
            date__month=month
        ).order_by('date')

        serializer = AttendanceRecordSerializer(attendance_records, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class DepartmentListView(APIView):
    def get(self, request):
        departments = Department.objects.all()
        serializer = DepartmentSerializer(departments, many=True)
        return Response(serializer.data)


class SalaryAvailableMonthsView(APIView):
    def get(self, request):
        months = Salary.objects.values_list('month', flat=True).distinct()
        return Response({'months': months})

class SalaryAvailableYearsView(APIView):
    def get(self, request):
        years = Salary.objects.values_list('year', flat=True).distinct()
        return Response({'years': years})


class UploadAttendanceFileView(APIView):
    def post(self, request, *args, **kwargs):

        print("hello")
        files = request.FILES.getlist('files')
        month = request.data.get('month')
        year = request.data.get('year')

        if not files or not month or not year:
            return Response({'error': 'Files, month, and year are required'}, status=status.HTTP_400_BAD_REQUEST)

        salary_responses = []

        for file in files:
            # Save file temporarily
            file_path = default_storage.save(f'temp/{file.name}', file)

            # Process file
            employee_info, attendance_data = self.process_attendance_file(file_path, month, year)

            print(f'post function attendance data {attendance_data}')

            # Clean up temporary file
            default_storage.delete(file_path)

            if employee_info is None or attendance_data is None:
                return Response({'error': f'Failed to process the attendance file {file.name}'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Process employee and attendance data
            employee = self.get_or_create_employee(employee_info)
            if employee is None:
                return Response({'error': f'Failed to process employee data for file {file.name}'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Add attendance records
            self.add_attendance_records(employee, attendance_data)

            # Call SalaryCalculationView to calculate salary
            salary_calculation_view = SalaryCalculationView()
            salary_request = request._request  # Get the original Django request
            salary_request.data = {
                'employee_code': employee.employee_code,
                'month': month,
                'year': year
            }
            salary_response = salary_calculation_view.post(salary_request)
            salary_responses.append({
                'employee_code': employee.employee_code,
                'response': salary_response.data,
                'status': salary_response.status_code
            })

        return Response({
            'message': 'All files processed successfully',
            'salary_responses': salary_responses
        }, status=status.HTTP_200_OK)


    def process_attendance_file(self, file_path, month, year):
        try:
            # Parse XML
            tree = ET.parse(file_path)
            root = tree.getroot()

            # Namespace dictionary for XML parsing
            namespaces = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

            # Extract rows and determine the maximum number of columns simultaneously
            rows = []
            max_cols = 0
            for row in root.findall(".//ss:Row", namespaces):
                cells = [cell.text for cell in row.findall("ss:Cell/ss:Data", namespaces)]
                rows.append(cells)
                max_cols = max(max_cols, len(cells))

            # Ensure all rows have the same number of columns
            for row in rows:
                row.extend([None] * (max_cols - len(row)))

            # Extract value of cell A2
            cell_A2 = rows[1][0] if len(rows) > 1 else 'No data'
            employee_info = self.parse_employee_info(cell_A2)

            # Convert to DataFrame
            df = pd.DataFrame(rows[1:], columns=rows[0])  # Assumes first row as header
            print(df)
            # Process DataFrame to extract attendance data for the given month and year
            attendance_data = self.extract_attendance_data(df, month, year)

            return employee_info, attendance_data

        except ET.ParseError as e:
            print(f"XML Parse Error: {e}")
        except Exception as e:
            print(f"Error: {e}")
        return None, None

    def parse_employee_info(self, cell_data):
        try:
            employee_code = self.extract_info(cell_data, "Mã nhân viên:", "Tên nhân viên:")
            employee_name = self.extract_info(cell_data, "Tên nhân viên:", "Bộ phận:")
            department_name = self.extract_info(cell_data, "Bộ phận:", None)
            return {
                'employee_code': employee_code,
                'employee_name': employee_name,
                'department_name': department_name
            }
        except Exception as e:
            print(f"Error parsing employee info: {e}")
            return None

    def extract_info(self, text, start_keyword, end_keyword):
        try:
            start = text.index(start_keyword) + len(start_keyword)
            if end_keyword:
                end = text.index(end_keyword, start)
                return text[start:end].strip()
            else:
                return text[start:].strip()
        except ValueError as e:
            print(f"Error extracting info between '{start_keyword}' and '{end_keyword}': {e}")
            return None



    def try_parsing_date(self, text):
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return pd.NaT


    def try_parsing_time(self, text):
        for fmt in ('%H:%M', '%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%H:%M:%S'):
            try:
                parsed_time = datetime.strptime(text, fmt)
                return parsed_time.strftime('%H:%M')
            except ValueError:
                continue
        return None



    def extract_attendance_data(self, df, month, year):
        # Assuming the first column contains the datetime information
        datetime_column = df.columns[0]

        df[datetime_column] = df[datetime_column].apply(lambda x: self.try_parsing_date(x) if pd.notnull(x) else pd.NaT)

        # Create start and end dates for the given month and year
        start_date = datetime(int(year), int(month), 1)
        end_date = (start_date + pd.DateOffset(months=1)) - pd.DateOffset(days=1)

        # Filter rows by the given month and year
        filtered_df = df[(df[datetime_column] >= start_date) & (df[datetime_column] <= end_date)]

        # Select relevant columns
        result_df = filtered_df.iloc[:, [0, 2, 3, 4, 5]]

        # Rename columns
        result_df.columns = ['Date', 'Morning In', 'Morning Out', 'Afternoon In', 'Afternoon Out']

        for col in ['Morning In', 'Morning Out', 'Afternoon In', 'Afternoon Out']:
            result_df.loc[:, col] = result_df[col].apply(lambda x: self.try_parsing_time(x) if pd.notnull(x) else None)

        # Convert DataFrame to dictionary
        attendance_data = result_df.to_dict(orient='records')

        return attendance_data

    def get_or_create_employee(self, employee_info):
        try:
            employee_code = employee_info['employee_code']
            employee_name = employee_info['employee_name']
            department_name = " ".join(employee_info['department_name'].split())

            # Kiểm tra xem nhân viên đã tồn tại chưa
            employee = Employee.objects.filter(employee_code=employee_code).first()
            if employee:
                return employee

            # Nếu nhân viên chưa tồn tại, tạo nhân viên mới
            employee = Employee.objects.create(employee_code=employee_code, name=employee_name)

            # Kiểm tra và tạo phòng ban nếu có thông tin
            if department_name:
                # Kiểm tra xem phòng ban đã tồn tại chưa (không phân biệt chữ hoa chữ thường)
                department = Department.objects.filter(name__iexact=department_name).first()
                if not department:
                    # Nếu phòng ban chưa tồn tại, tạo phòng ban mới
                    department = Department.objects.create(name=department_name)

                # Thêm nhân viên vào phòng ban nếu chưa tồn tại
                if not EmployeeDepartment.objects.filter(employee=employee, department=department).exists():
                    EmployeeDepartment.objects.create(employee=employee, department=department)

            return employee
        except Exception as e:
            print(f"Error getting or creating employee: {e}")
            return None

    def add_attendance_records(self, employee, attendance_data):
        try:
            for record in attendance_data:
                attendance_record, created = AttendanceRecord.objects.update_or_create(
                    employee=employee,
                    date=record['Date'],
                    defaults={
                        'morning_clock_in': record['Morning In'],
                        'morning_clock_out': record['Morning Out'],
                        'afternoon_clock_in': record['Afternoon In'],
                        'afternoon_clock_out': record['Afternoon Out']
                    }
                )
        except Exception as e:
            print(f"Error adding attendance records: {e}")




class SalaryCalculationView(APIView):
    def post(self, request, *args, **kwargs):
        employee_code = request.data.get('employee_code')
        month = request.data.get('month')
        year = request.data.get('year')


        if not employee_code or not month or not year:
            return Response({'error': 'Employee ID, month, and year are required'}, status=status.HTTP_400_BAD_REQUEST)

        employee = Employee.objects.filter(employee_code=employee_code).first()
        if not employee:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        # Gọi hàm tính lương
        salary_data, errors = self.calculate_salary(employee, month, year)

        if errors:
            Salary.objects.update_or_create(
                employee=employee,
                month=month,
                year=year,
                defaults={
                    'basic_days_after_holidays': None,
                    'basic_hours_after_holidays': None,
                    'actual_work_hours': None,
                    'worked_days': None,
                    'penalty_hours': None,
                    'worked_day_off_days': None,
                    'sunday_hours': None,
                    'holiday_hours': None,
                    'worked_holiday_hours': None,
                    'average_hours_per_day': None,
                    'worked_day_off_hours': None,
                    'overtime_hours': None,
                    'total_hours': None,
                    'salary_amount': None
                }
            )




            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        # salary_data['employee'] = employee

        # Tạo hoặc cập nhật bản ghi Salary
        salary_instance, created = Salary.objects.update_or_create(
            employee=employee,
            month=month,
            year=year,
            defaults=salary_data
        )

        return Response(SalarySerializer(salary_instance).data, status=status.HTTP_201_CREATED)

    def calculate_salary(self, employee, month, year):
        errors = []

        # Lấy tất cả các bản ghi chấm công trong tháng và năm cho nhân viên
        attendance_records = AttendanceRecord.objects.filter(
            employee=employee,
            date__year=year,
            date__month=month
        )

        config = Configuration.objects.first()
        if not config:
            return {}, [{'error': 'Configuration not found'}]

        holidays = Holiday.objects.filter(month=month)
        total_holidays = holidays.count()

        basic_day_after_holidays = config.basic_days - total_holidays
        basic_hours_after_holidays = basic_day_after_holidays * config.basic_hours_per_day

        holiday_hours = total_holidays * config.basic_hours_per_day
        worked_holiday_hours = 0

        # Khởi tạo các biến lưu trữ giờ làm việc
        actual_work_hours = 0
        worked_days = 0
        penalty_hours = 0
        worked_day_off_days = 0
        sunday_hours = 0
        working_sunday_days = 0

        average_hours_per_day = 0
        worked_day_off_hours = 0
        overtime_hours = 0
        total_hours = 0
        salary_amount = 0

        for record in attendance_records:
            morning_duration, afternoon_duration, daily_hours, day_errors = self.calculate_daily_hours(record)
            if day_errors:
                errors.append({
                    'date': record.date,
                    'errors': day_errors
                })
            else:
                actual_work_hours += daily_hours

                if morning_duration > 0:
                    worked_days += 3 / 8
                if afternoon_duration > 0:
                    worked_days += 5 / 8
                print(worked_days)
                # Kiểm tra và tính toán giờ làm việc vào Chủ nhật
                if record.date.weekday() == calendar.SUNDAY:
                    sunday_hours += daily_hours
                    working_sunday_days += 1

                # Kiểm tra và tính toán giờ làm việc vào ngày lễ
                if holidays.filter(day=record.date.day).exists():
                    worked_holiday_hours += daily_hours
        print(worked_days)

        average_hours_per_day = actual_work_hours / worked_days if worked_days else 0
        worked_day_off_days = worked_days - basic_day_after_holidays

        if worked_day_off_days >= 0:
            worked_day_off_hours = (average_hours_per_day * worked_day_off_days)
        else:
            sunday_hours = 0
            worked_day_off_days = 0
            worked_day_off_hours = 0

        overtime_hours = max(0, actual_work_hours - basic_hours_after_holidays - worked_day_off_hours)


        if actual_work_hours < basic_hours_after_holidays:
            total_hour_none_bonus = max(0, actual_work_hours - worked_holiday_hours - worked_day_off_hours - overtime_hours)
        else:
            total_hour_none_bonus = basic_hours_after_holidays

        print(f'none bonus: {total_hour_none_bonus}')

        # Tính tổng số giờ
        total_hours = (total_hour_none_bonus
                       + sunday_hours
                       + holiday_hours
                       + worked_holiday_hours * config.holiday_multiplier
                       + worked_day_off_hours * config.worked_day_off_multiplier
                       + overtime_hours * config.overtime_multiplier - penalty_hours)

        print(f'total hour {total_hours}')

        # Tính số tiền lương
        total_hours = math.ceil(total_hours)
        salary_amount = total_hours * config.hourly_wage

        salary_data = {
            'employee': employee,
            'month': month,
            'year': year,
            'basic_days_after_holidays': basic_day_after_holidays,
            'basic_hours_after_holidays': basic_hours_after_holidays,
            'actual_work_hours': actual_work_hours,
            'worked_days': worked_days,
            'penalty_hours': penalty_hours,
            'worked_day_off_days': worked_day_off_days,
            'sunday_hours': sunday_hours,
            'holiday_hours': holiday_hours,
            'worked_holiday_hours': worked_holiday_hours,
            'average_hours_per_day': average_hours_per_day,
            'worked_day_off_hours': worked_day_off_hours,
            'overtime_hours': overtime_hours,
            'total_hours': total_hours,
            'salary_amount': salary_amount
        }

        return salary_data, errors

    def calculate_daily_hours(self, record):
        errors = []
        total_hours = 0
        morning_duration = 0
        afternoon_duration = 0

        def time_difference(start_time, end_time):
            # Sử dụng một ngày cụ thể để tính toán sự khác biệt
            start_datetime = datetime.combine(datetime.today(), start_time)
            end_datetime = datetime.combine(datetime.today(), end_time)
            return (end_datetime - start_datetime).total_seconds() / 3600

        if record.morning_clock_in and record.morning_clock_out:
            morning_duration += time_difference(record.morning_clock_in, record.morning_clock_out)
            total_hours += morning_duration
        elif record.morning_clock_in or record.morning_clock_out:
            if not record.morning_clock_in:
                errors.append({
                    'error_type': 'morning_clock_in',
                    'message': 'Missing morning clock-in time'
                })
            if not record.morning_clock_out:
                errors.append({
                    'error_type': 'morning_clock_out',
                    'message': 'Missing morning clock-out time'
                })

        if record.afternoon_clock_in and record.afternoon_clock_out:
            afternoon_duration += time_difference(record.afternoon_clock_in, record.afternoon_clock_out)
            total_hours += afternoon_duration
        elif record.afternoon_clock_in or record.afternoon_clock_out:
            if not record.afternoon_clock_in:
                errors.append({
                    'error_type': 'afternoon_clock_in',
                    'message': 'Missing afternoon clock-in time'
                })
            if not record.afternoon_clock_out:
                errors.append({
                    'error_type': 'afternoon_clock_out',
                    'message': 'Missing afternoon clock-out time'
                })

        return morning_duration, afternoon_duration, total_hours, errors






def export_employee_salary_report(request, employee_code, month, year):
    try:
        # Lấy dữ liệu chấm công và lương từ database
        employee = Employee.objects.get(employee_code=employee_code)
        attendance_records = AttendanceRecord.objects.filter(employee=employee, date__year=year, date__month=month)
        salary_record = Salary.objects.get(employee=employee, month=month, year=year)

        # Tạo workbook và các sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Employee Report'


        if salary_record.salary_amount == 0 or salary_record.salary_amount is None:
            salary_calculation_view = SalaryCalculationView()
            _, errors = salary_calculation_view.calculate_salary(employee, month, year)
        else:

            errors = []


        print(errors)

        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 17
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['E'].width = 17
        sheet.column_dimensions['F'].width = 17
        sheet.column_dimensions['G'].width = 17

        # Ghi tiêu đề
        sheet['A1'] = 'Bảng chi tiết chấm công'
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:F1')

        # Ghi thông tin nhân viên
        sheet['A2'] = f'Mã nhân viên: {employee.employee_code} Tên nhân viên: {employee.name} Bộ phận: {"".join([dept.name for dept in employee.departments.all()])}'
        sheet.merge_cells('A2:F2')
        sheet['A2'].font = Font(size=12)
        sheet['A2'].alignment = Alignment(horizontal='center')

        # Định dạng border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Ghi bảng chấm công
        headers = ['Ngày', 'Thứ', 'Giờ vào buổi sáng', 'Giờ ra buổi sáng', 'Giờ vào buổi chiều', 'Giờ ra buổi chiều']
        header_row = sheet.append(headers)
        for cell in sheet[3]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # for record in attendance_records:
        #     day_of_week = calendar.day_name[record.date.weekday()]
        #     row = [
        #         record.date.strftime('%d/%m/%Y'),
        #         day_of_week,
        #         record.morning_clock_in.strftime('%H:%M') if record.morning_clock_in else 'N/A',
        #         record.morning_clock_out.strftime('%H:%M') if record.morning_clock_out else 'N/A',
        #         record.afternoon_clock_in.strftime('%H:%M') if record.afternoon_clock_in else 'N/A',
        #         record.afternoon_clock_out.strftime('%H:%M') if record.afternoon_clock_out else 'N/A'
        #     ]
        #     sheet.append(row)
        #     for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=6):
        #         for c in cell:
        #             c.alignment = Alignment(horizontal='center')
        #             c.border = thin_border

        def get_error_message(date, error_type):
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()  # Chuyển đổi ngày từ chuỗi về dạng datetime.date
            for error in errors:
                if error['date'] == date_obj:
                    for err in error['errors']:
                        if err['error_type'] == error_type:
                            return "Missing time"
            return "N/A"

        for record in attendance_records:
            day_of_week = calendar.day_name[record.date.weekday()]
            row = [
                record.date.strftime('%d/%m/%Y'),
                day_of_week,
                record.morning_clock_in.strftime('%H:%M') if record.morning_clock_in else get_error_message(record.date.strftime('%Y-%m-%d'), "morning_clock_in"),
                record.morning_clock_out.strftime('%H:%M') if record.morning_clock_out else get_error_message(record.date.strftime('%Y-%m-%d'), "morning_clock_out"),
                record.afternoon_clock_in.strftime('%H:%M') if record.afternoon_clock_in else get_error_message(record.date.strftime('%Y-%m-%d'), "afternoon_clock_in"),
                record.afternoon_clock_out.strftime('%H:%M') if record.afternoon_clock_out else get_error_message(record.date.strftime('%Y-%m-%d'), "afternoon_clock_out")
            ]
            sheet.append(row)
            for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=6):
                for c in cell:
                    c.alignment = Alignment(horizontal='center')
                    c.border = thin_border







        # Thêm dòng trống giữa bảng chấm công và chi tiết lương
        sheet.append([])

        # Ghi chi tiết lương
        sheet.append(['Chi tiết lương'])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
        sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(horizontal='left')
        salary_fields = [
            ('Tháng', salary_record.month),
            ('Năm', salary_record.year),
            ('Ngày cơ bản sau lễ', salary_record.basic_days_after_holidays),
            ('Giờ cơ bản sau lễ', salary_record.basic_hours_after_holidays),
            ('Giờ làm việc thực tế', salary_record.actual_work_hours),
            ('Ngày đi làm', salary_record.worked_days),
            ('Giờ vi phạm', salary_record.penalty_hours),
            ('Ngày nghỉ đi làm', salary_record.worked_day_off_days),
            ('Giờ làm Chủ nhật', salary_record.sunday_hours),
            ('Giờ lễ', salary_record.holiday_hours),
            ('Giờ làm ngày lễ', salary_record.worked_holiday_hours),
            ('Giờ trung bình mỗi ngày', salary_record.average_hours_per_day),
            ('Giờ nghỉ đi làm', salary_record.worked_day_off_hours),
            ('Giờ làm thêm', salary_record.overtime_hours),
            ('Tổng giờ', salary_record.total_hours),
            ('Số tiền lương', salary_record.salary_amount)
        ]

        for field_name, field_value in salary_fields:
            row = sheet.append([field_name, field_value])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
            sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(horizontal='left')
            for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=2):
                for c in cell:
                    c.border = thin_border



        # Tạo response và đính kèm file Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={employee_code}_Report_{month}_{year}.xlsx'
        workbook.save(response)
        return response

    except Employee.DoesNotExist:
        return HttpResponse(status=404, content='Employee not found')
    except Salary.DoesNotExist:
        return HttpResponse(status=404, content='Salary record not found')

from django.db.models import Q

def export_salary_summary(request):
    try:
        # # Lấy dữ liệu lương từ database
        # salaries = Salary.objects.filter(month=month, year=year)
        #
        # if department:
        #     salaries = salaries.filter(employee__departments__id=department)

        month = request.GET.get('month')  # Lấy tham số month
        year = request.GET.get('year')  # Lấy tham số year
        department = request.GET.get('department')  # Lấy tham số department

        # Bắt đầu lọc dữ liệu
        salaries = Salary.objects.all()
        if month:
            salaries = salaries.filter(month=month)
        if year:
            salaries = salaries.filter(year=year)
        if department:
            salaries = salaries.filter(employee__departments__id=department)

        # Thay đổi tiêu đề dựa trên các tham số truyền vào
        title_parts = []
        if month:
            title_parts.append(f'tháng {month}')
        if year:
            title_parts.append(f'năm {year}')
        title = 'Báo cáo lương ' + ' '.join(title_parts) if title_parts else 'Báo cáo lương tổng hợp'



        # Tạo workbook và các sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Monthly Salary Summary'

        # Định dạng các cột
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 27
        sheet.column_dimensions['C'].width = 35
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 20

        # Ghi tiêu đề chính
        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = title
        # f'Báo cáo lương tháng {month}/{year}'
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Thêm dòng trống
        sheet.append([])

        # Ghi tiêu đề các cột
        headers = ['Mã nhân viên', 'Tên nhân viên', 'Bộ phận', 'Tháng', 'Năm', 'Lương']
        sheet.append(headers)
        for cell in sheet[3]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Định dạng border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Ghi dữ liệu lương vào sheet
        for salary in salaries:
            departments = ", ".join([dept.name for dept in salary.employee.departments.all()])
            salary_amount = salary.salary_amount if salary.salary_amount is not None else 'Thiếu giờ check-in'
            if salary_amount == 0:
                salary_amount = 0

            row = [
                salary.employee.employee_code,
                salary.employee.name,
                departments,
                salary.month,
                salary.year,
                salary_amount
            ]
            sheet.append(row)
            for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=6):
                for c in cell:
                    c.alignment = Alignment(horizontal='center')
                    c.border = thin_border

        # Thêm border cho toàn bộ bảng
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border

        # Tạo response và đính kèm file Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Salary_Summary_{month}_{year}.xlsx'
        workbook.save(response)
        return response

    except Exception as e:
        return HttpResponse(status=500, content=str(e))
